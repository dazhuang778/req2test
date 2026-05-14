from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import TestCase

_HEADERS = ["用例编号", "标题", "前置条件", "步骤", "预期结果"]
_COL_WIDTHS = [15, 30, 40, 40, 40]
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_WRAP_COLS = {4, 5}  # 步骤、预期结果（1-indexed）


def export_to_excel(test_cases: list[TestCase], output_path: str | Path) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    ws.append(_HEADERS)
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        ws.column_dimensions[cell.column_letter].width = width

    for case in test_cases:
        row = [
            case.id,
            case.title,
            case.preconditions,
            case.steps,
            case.expected_result,
        ]
        ws.append(row)
        data_row = ws.max_row
        for col_idx in _WRAP_COLS:
            ws.cell(row=data_row, column=col_idx).alignment = Alignment(wrap_text=True)

    wb.save(output_path)
